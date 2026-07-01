from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
OUT = Path(__file__).resolve().parent / "data.js"


def clean(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def slug(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return value or "item"


def parse_unit_name(raw: str) -> tuple[str, str, str]:
    raw = clean(raw)
    match = re.match(r"(.+?)\s*\(([^)]+)\)", raw)
    if match:
        city = clean(match.group(1)).title()
        state = clean(match.group(2)).upper()
    else:
        parts = raw.rsplit(" ", 1)
        city = clean(parts[0] if len(parts) > 1 and len(parts[1]) == 2 else raw).title()
        state = clean(parts[1]).upper() if len(parts) > 1 and len(parts[1]) == 2 else ""
    return raw, city, state


def parse_roadmap_file(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb["RoadMap de Implantação"]
    raw_name, city, state = parse_unit_name(ws["B2"].value)
    unit_id = slug(f"{city}-{state}" if state else city)
    unit = {
        "id": unit_id,
        "name": raw_name,
        "city": city,
        "state": state,
        "franchisee": clean(ws["B3"].value),
        "openingDate": clean(ws["B4"].value),
        "sourceFile": path.name,
        "tasks": [],
        "purchases": [],
    }

    current_phase = ""
    for row in range(6, ws.max_row + 1):
        item = clean(ws.cell(row, 1).value)
        process = clean(ws.cell(row, 2).value)
        status = clean(ws.cell(row, 3).value)
        deadline = clean(ws.cell(row, 4).value)
        actual = clean(ws.cell(row, 5).value)
        notes = clean(ws.cell(row, 6).value)

        if not any([item, process, status, deadline, actual, notes]):
            continue
        if item.upper() in {"ITEM", "PROCESSO / ETAPA"} or process.upper() == "PROCESSO / ETAPA":
            continue
        if item and not process and not status:
            current_phase = item.upper()
            continue
        if not process:
            continue

        unit["tasks"].append(
            {
                "id": f"{unit_id}-{slug(str(item))}-{slug(process)}",
                "item": item,
                "phase": current_phase or "SEM FASE",
                "process": process,
                "status": status or "Sem status",
                "deadline": deadline,
                "actualDate": actual,
                "notes": notes,
            }
        )

    ws = wb["Checklist de Compras"]
    for row in range(3, ws.max_row + 1):
        item = clean(ws.cell(row, 1).value)
        if not item or item.upper() == "ITEM":
            continue
        unit["purchases"].append(
            {
                "id": f"{unit_id}-compra-{slug(item)}",
                "item": item,
                "status": clean(ws.cell(row, 2).value) or "Sem status",
                "notes": clean(ws.cell(row, 3).value),
            }
        )

    return unit


def parse_accreditation(path: Path, known_units: list[dict]) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    known_by_city = {slug(unit["city"]): unit["id"] for unit in known_units}
    columns = []
    for col in range(3, ws.max_column + 1):
        label = clean(ws.cell(2, col).value)
        if not label:
            continue
        normalized = slug(label.replace("ANANIDEUA", "ANANINDEUA"))
        unit_id = known_by_city.get(normalized, normalized)
        columns.append(
            {
                "column": col,
                "id": unit_id,
                "name": label,
                "owner": clean(ws.cell(3, col).value),
            }
        )

    procedures = []
    current_group = ""
    for row in range(4, ws.max_row + 1):
        name = clean(ws.cell(row, 1).value)
        if not name:
            continue
        if name.upper().startswith("GRUPO"):
            current_group = name.upper()
            continue

        statuses = {}
        for col in columns:
            status = clean(ws.cell(row, col["column"]).value).upper()
            if status:
                statuses[col["id"]] = status
        if statuses:
            procedures.append(
                {
                    "id": f"{slug(current_group)}-{slug(name)}",
                    "group": current_group or "SEM GRUPO",
                    "name": name,
                    "statuses": statuses,
                }
            )

    return {
        "sourceFile": path.name,
        "units": [{"id": c["id"], "name": c["name"], "owner": c["owner"]} for c in columns],
        "procedures": procedures,
    }


def summarize(units: list[dict], accreditation: dict) -> dict:
    task_status = Counter(task["status"] for unit in units for task in unit["tasks"])
    purchase_status = Counter(item["status"] for unit in units for item in unit["purchases"])
    accreditation_status = Counter(
        status
        for procedure in accreditation["procedures"]
        for status in procedure["statuses"].values()
    )
    return {
        "taskStatus": dict(task_status),
        "purchaseStatus": dict(purchase_status),
        "accreditationStatus": dict(accreditation_status),
    }


def main() -> None:
    roadmap_files = sorted(
        path
        for path in ROOT.glob("*.xlsx")
        if "CREDENCIAMENTOS" not in path.name.upper() and not path.name.startswith("~$")
    )
    units = [parse_roadmap_file(path) for path in roadmap_files]
    accreditation_file = next(path for path in ROOT.glob("CREDENCIAMENTOS*.xlsx"))
    accreditation = parse_accreditation(accreditation_file, units)

    model_tasks = []
    seen_tasks = set()
    for unit in units:
        for task in unit["tasks"]:
            key = slug(task["process"])
            if key in seen_tasks:
                continue
            seen_tasks.add(key)
            model_tasks.append({"phase": task["phase"], "item": task["item"], "process": task["process"]})

    purchase_items = sorted({purchase["item"] for unit in units for purchase in unit["purchases"]})
    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFiles": [path.name for path in roadmap_files] + [accreditation_file.name],
        "units": units,
        "accreditation": accreditation,
        "modelTasks": model_tasks,
        "purchaseItems": purchase_items,
        "summary": summarize(units, accreditation),
    }

    OUT.write_text(
        "window.FRANCHISE_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
